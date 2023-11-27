import os.path
import zipfile
from io import TextIOWrapper

from pypdf import PdfReader
from openpyxl import load_workbook
import csv

CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIR = os.path.dirname(CURRENT_FILE)

downloads = os.path.join(CURRENT_DIR, "downloads")
resources = os.path.join(CURRENT_DIR, "resources")


def create_archive():
    file_dir = os.listdir(downloads)
    archive_name = os.path.join(resources, 'test.zip')

    if not os.path.exists(resources):
        os.mkdir(resources)
    if os.path.isfile(archive_name):
        os.remove(archive_name)

    with zipfile.ZipFile(archive_name, mode='w', compression=zipfile.ZIP_DEFLATED) as archive:
        for file in file_dir:
            archive.write(os.path.join(downloads, file), file)

    return archive_name


def test_archive():
    archive_name = create_archive()
    print(archive_name)

    with zipfile.ZipFile(archive_name) as zip_file:
        print(zip_file.namelist())
        with zip_file.open('example.pdf') as pdf:
            reader = PdfReader(pdf)
            text = reader.pages[0].extract_text()
            assert 'Пример PDF файла' in text
        with zip_file.open('import_ou_xlsx.xlsx') as xlsx:
            workbook = load_workbook(xlsx)
            sheet = workbook.active
            first_last = sheet.cell(row=7, column=1).value
            second_last = sheet.cell(row=7, column=2).value
            third_last = sheet.cell(row=7, column=3).value
            assert first_last == "OU006"
            assert second_last == "OU005"
            assert third_last == "Закупки"

        with zip_file.open('file.csv', 'r') as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8')))
            assert 'superman' == csvreader[2][0]
            assert 'flies through the air' == csvreader[2][1]
